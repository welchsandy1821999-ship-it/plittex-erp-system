import openpyxl

def check_sklad2(file_path):
    wb = openpyxl.load_workbook(file_path, read_only=True, data_only=True)
    print(f"Sheets: {wb.sheetnames}")
    for sheet_name in wb.sheetnames:
        sheet = wb[sheet_name]
        print(f"Sheet '{sheet_name}': max_row={sheet.max_row}, max_col={sheet.max_column}")
        for i, row in enumerate(sheet.iter_rows(max_row=5), 1):
            vals = [str(cell.value) if cell.value is not None else "" for cell in row]
            print(f"Row {i}: {vals}")

check_sklad2(r'c:\Users\Пользователь\Desktop\plittex-erp\склад2.xlsx')
