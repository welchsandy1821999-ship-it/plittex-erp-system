import openpyxl
import re
import csv
import json
import sys
from collections import defaultdict

# Helper to normalize names for comparison
def norm(s):
    if not s: return ""
    return re.sub(r'[^\w]', '', str(s).lower()).replace('x', 'х')

def clean_price(val):
    if val is None: return None
    if isinstance(val, (int, float)): return float(val)
    if isinstance(val, str):
        s = re.sub(r'[^\d]', '', val)
        try: return float(s)
        except: return None
    return None

def get_grade_info(full_name):
    grade_mult = 1.0
    grade_label = "1 сорт"
    if "2 сорт" in full_name:
        grade_mult = 0.5
        grade_label = "2 сорт"
    elif "Экспериментальная" in full_name or "Эксперементальная" in full_name:
        grade_mult = 0.7
        grade_label = "Экспериментальная"
    
    # Base name for price matching
    base_name = re.sub(r'2 сорт|Экспериментальная|Эксперементальная', '', full_name).strip()
    return base_name, grade_mult, grade_label

def generate_csv():
    price_file = r'c:\Users\Пользователь\Desktop\plittex-erp\Price_март2026.xlsx'
    sklad_file = r'c:\Users\Пользователь\Desktop\plittex-erp\склад2.xlsx'
    output_file = r'c:\Users\Пользователь\Desktop\plittex-erp\ревизия_склада_март2026.csv'
    
    # 1. Harvest ALL Prices from Price List (General)
    prices = {} # {Key: Price}
    try:
        wb_p = openpyxl.load_workbook(price_file, read_only=True, data_only=True)
        sheet_p = wb_p.active
        for row in sheet_p.iter_rows(min_row=1):
            row_vals = [cell.value for cell in row]
            if len(row_vals) < 9: continue
            name = str(row_vals[1]).strip().lower() if row_vals[1] and len(str(row_vals[1])) > 3 else ""
            dims = str(row_vals[3]).strip().lower() if row_vals[3] else ""
            p_grey = clean_price(row_vals[6])
            if dims and p_grey:
                key = re.sub(r'[^\w]', '', name + dims).replace('x', 'х')
                prices[key] = p_grey
    except: pass

    # Manual dictionary for Classico/Uriko/Plates
    PRICES_MANUAL = {
        "2К0,4": (1200.0, 1400.0), # Standard 40mm
        "2К0,6": (1400.0, 1500.0), # Standard 60mm
        "2К4": (1310.0, 1500.0),   # Granit 40mm
        "2К6": (1550.0, 1650.0),   # Granit 60mm
        "Урико": (1200.0, 1400.0), 
        "2П4": (1500.0, 1750.0),   
        "2П6": (1650.0, 1950.0),   
        "3.П.8": (2250.0, 2650.0), 
        "1.ПР.8": (2250.0, 2650.0), 
    }

    # 2. Process склад2.xlsx
    inventory = defaultdict(float)
    try:
        wb_s = openpyxl.load_workbook(sklad_file, read_only=True, data_only=True)
        ws_s = wb_s.worksheets[0]
        for row in ws_s.iter_rows(min_row=2):
            name_val = row[6].value # Col G
            qty_val = row[7].value # Col H
            if name_val:
                inventory[str(name_val).strip()] += float(qty_val or 0)
    except: pass

    fallbacks = {
        "бордюрдорожный": 700.0,
        "поребрик": 680.0,
        "ситимикс": 950.0,
        "паркет": 1000.0,
        "плита": 900.0
    }

    final_rows = []
    for name, qty in inventory.items():
        base_name, grade_mult, grade_label = get_grade_info(name)
        lower_name = base_name.lower()
        price = 0.0
        
        for pk, pv in PRICES_MANUAL.items():
            if pk.lower() in lower_name:
                price = pv[0] if any(s in lower_name for s in ["серый", "серая"]) else pv[1]
                break
        
        if price == 0:
            match_key = re.sub(r'[^\w]', '', lower_name).replace('x', 'х')
            price = prices.get(match_key, 0.0)
            
        if price == 0:
            for fk, fv in fallbacks.items():
                if fk in re.sub(r'[^\w]', '', lower_name):
                    price = fv if any(s in lower_name for s in ["серый", "серая"]) else fv + 150.0
                    break
        
        calc_price = price * grade_mult
        final_rows.append([name, grade_label, f"{qty:.2f}", f"{price:.2f}", f"{calc_price:.2f}", f"{calc_price * qty:.2f}"])

    final_rows.sort(key=lambda x: x[0])

    # Write CSV
    with open(output_file, 'w', newline='', encoding='utf-8-sig') as f:
        writer = csv.writer(f, delimiter=';')
        writer.writerow(["Название товара", "Сорт", "Кол-во", "Базовая цена", "Цена за ед.", "Итого"])
        writer.writerows(final_rows)
        grand_total = sum(float(r[5]) for r in final_rows)
        writer.writerow([])
        writer.writerow(["ИТОГО ПО СКЛАДУ", "", "", "", "", f"{grand_total:.2f}"])

    print(f"CSV generated: {output_file}")

if __name__ == "__main__":
    generate_csv()
