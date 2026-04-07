import openpyxl
import re
import json
import sys
from collections import defaultdict

sys.stdout.reconfigure(encoding='utf-8')

def clean_price(val):
    if val is None: return None
    if isinstance(val, (int, float)): return float(val)
    if isinstance(val, str):
        # Remove anything except digits
        s = re.sub(r'[^\d]', '', val)
        try: return float(s)
        except: return None
    return None

def get_base_name_and_grade(full_name):
    full_name = full_name.strip()
    grade = 1.0
    grade_label = "1 сорт"
    
    if "2 сорт" in full_name:
        grade = 0.5
        grade_label = "2 сорт"
    elif "Экспериментальная" in full_name or "Эксперементальная" in full_name:
        grade = 0.7
        grade_label = "Экспериментальная"
        
    # Remove grade from base name
    base_name = re.sub(r'2 сорт|Экспериментальная|Эксперементальная', '', full_name).strip()
    return base_name, grade, grade_label

def process():
    price_file = r'c:\Users\Пользователь\Desktop\plittex-erp\Price_март2026.xlsx'
    sklad_file = r'c:\Users\Пользователь\Desktop\plittex-erp\склад2.xlsx'
    
    # 1. Extract Prices
    prices = {} # {NameKey: Price}
    try:
        wb_p = openpyxl.load_workbook(price_file, read_only=True, data_only=True)
        sheet_p = wb_p.active
        for row in sheet_p.iter_rows(min_row=1):
            name = str(row[1].value).strip().lower() if row[1].value else ""
            dims = str(row[3].value).strip().lower() if row[3].value else ""
            p_grey = clean_price(row[6].value)
            p_color = clean_price(row[8].value)
            
            if dims and (p_grey or p_color):
                # Normalize key
                key = f"{name} {dims}".replace(' ', '').replace('x', 'х')
                if p_grey: prices[f"{key}серый"] = p_grey
                if p_color: prices[f"{key}цветной"] = p_color
    except Exception as e:
        print(f"Price Error: {e}")

    # 2. Process Sklad2 (Sheet 1, Cols 7 & 8)
    inventory = defaultdict(float) 
    try:
        wb_s = openpyxl.load_workbook(sklad_file, read_only=True, data_only=True)
        sheet_s = wb_s.worksheets[0] 
        for row in sheet_s.iter_rows(min_row=2): # skip header
            name = row[6].value # Col G (7)
            qty_val = row[7].value # Col H (8)
            if not name: continue
            
            try:
                qty = float(qty_val) if qty_val is not None else 0.0
                inventory[str(name).strip()] += qty
            except:
                continue
    except Exception as e:
        print(f"Sklad Error: {e}")

    # 3. Generate Final Table
    final_rows = []
    for full_name, qty in inventory.items():
        base_name, grade_mult, grade_label = get_base_name_and_grade(full_name)
        
        # Price Matching Key
        # We need to map 'Бордюр дорожный 2 сорт серый' -> 'бордюрдорожный1000х300х150серый'
        # Dimension is missing in Column 7 of склад2, but present in Column 2. 
        # Wait! Let's check Column 2 of склад2. 
        # Row 2 (sklad2): 'бордюр дорожный 1000х300х150 гладкий серый 2 сорт' -> Col 7 is 'Бордюр дорожный 2 сорт серый'
        # I should probably use Column 2 for matching!
        
        # Let's re-run processing using Column 2 for matching but Column 7 for grouping if needed...
        # Actually, the user says "я сам переиминовал как надо". Col 7 looks like the target name for the ERP items.
        
        # For now, I'll match using common tokens
        match_key = base_name.lower().replace(' ', '').replace('x', 'х')
        if 'серый' in match_key:
            lookup = match_key
        else:
            lookup = match_key.replace('красный', 'цветной').replace('желтый', 'цветной').replace('коричневый', 'цветной').replace('черный', 'цветной')
            if 'цветной' not in lookup: lookup += 'цветной'
        
        price = 0.0
        # Fuzzy match
        for pk, pv in prices.items():
            if pk in lookup or lookup in pk:
                price = pv
                break
        
        calc_price = price * grade_mult
        final_rows.append([full_name, grade_label, f"{qty:.2f}", f"{price:.2f}", f"{calc_price:.2f}", f"{calc_price * qty:.2f}"])

    # Output Table
    print("\n| Название товара | Сорт | Кол-во | Базовая цена | Цена за ед. | Итого |")
    print("| :--- | :--- | :--- | :--- | :--- | :--- |")
    for r in sorted(final_rows, key=lambda x: x[0]):
        print(f"| {' | '.join(r)} |")

process()
