import openpyxl
import re
import json
import sys
from collections import defaultdict

sys.stdout.reconfigure(encoding='utf-8')

def clean_price(val):
    if val is None: return None
    if isinstance(val, (int, float)): return float(val)
    if isinstance(val, str):
        # Remove anything except digits
        s = re.sub(r'[^\d]', '', val)
        try: return float(s)
        except: return None
    return None

def get_base_name_and_grade(full_name):
    full_name = full_name.strip()
    grade = 1.0
    grade_label = "1 сорт"
    
    if "2 сорт" in full_name:
        grade = 0.5
        grade_label = "2 сорт"
    elif "Экспериментальная" in full_name or "Эксперементальная" in full_name:
        grade = 0.7
        grade_label = "Экспериментальная"
        
    # Remove grade from base name
    base_name = re.sub(r'2 сорт|Экспериментальная|Эксперементальная', '', full_name).strip()
    return base_name, grade, grade_label

def process():
    price_file = r'c:\Users\Пользователь\Desktop\plittex-erp\Price_март2026.xlsx'
    sklad_file = r'c:\Users\Пользователь\Desktop\plittex-erp\склад2.xlsx'
    
    # 1. Extract Prices
    prices = {} # {NameKey: Price}
    try:
        wb_p = openpyxl.load_workbook(price_file, read_only=True, data_only=True)
        sheet_p = wb_p.active
        for row in sheet_p.iter_rows(min_row=1):
            name = str(row[1].value).strip().lower() if row[1].value else ""
            dims = str(row[3].value).strip().lower() if row[3].value else ""
            p_grey = clean_price(row[6].value)
            p_color = clean_price(row[8].value)
            
            if dims and (p_grey or p_color):
                key = f"{name} {dims}".replace(' ', '').replace('x', 'х')
                if p_grey: prices[f"{key}серый"] = p_grey
                if p_color: prices[f"{key}цветной"] = p_color
        print(f"DEBUG: Extracted {len(prices)} prices from Excel.")
    except Exception as e:
        print(f"Price Error: {e}")

    # 2. Process Sklad2
    inventory = defaultdict(float) 
    try:
        wb_s = openpyxl.load_workbook(sklad_file, read_only=True, data_only=True)
        sheet_s = wb_s.active
        row_count = 0
        for row in sheet_s.iter_rows(min_row=1):
            name = row[0].value
            if not name: continue
            
            qty_raw = str(row[1].value) if row[1].value else "0"
            qty_match = re.search(r'[\d,.]+', qty_raw)
            qty = float(qty_match.group().replace(',', '.')) if qty_match else 0.0
            
            inventory[name.strip()] += qty
            row_count += 1
        print(f"DEBUG: Read {row_count} rows from склад2.xlsx, unique items: {len(inventory)}")
    except Exception as e:
        print(f"Sklad Error: {e}")

    # 3. Generate Final Table
    final_rows = []
    for full_name, qty in inventory.items():
        base_name, grade_mult, grade_label = get_base_name_and_grade(full_name)
        
        # Price Matching
        search_key = base_name.lower().replace(' ', '').replace('x', 'х')
        if 'серый' in search_key:
            lookup = search_key # Should already have 'серый'
        else:
            # If color, we assume it's one of the non-grey colors (p_color)
            lookup = search_key.replace('красный', 'цветной').replace('желтый', 'цветной').replace('коричневый', 'цветной').replace('черный', 'цветной')
            if 'цветной' not in lookup:
                lookup += 'цветной'
        
        price = prices.get(lookup, 0.0)
        
        # Fuzzy match if exact fails
        if price == 0 and len(prices) > 0:
            for pk, pv in prices.items():
                if pk in lookup or lookup in pk:
                    price = pv
                    break
        
        calc_price = price * grade_mult
        final_rows.append([full_name, grade_label, f"{qty:.2f}", f"{price:.2f}", f"{calc_price:.2f}", f"{calc_price * qty:.2f}"])

    # Print Table
    print("\n| Название товара | Сорт | Кол-во | Базовая цена | Цена за ед. | Итого |")
    print("| :--- | :--- | :--- | :--- | :--- | :--- |")
    for r in sorted(final_rows, key=lambda x: x[0]):
        print(f"| {' | '.join(r)} |")

process()
