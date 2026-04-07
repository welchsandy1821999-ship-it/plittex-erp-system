import openpyxl
import re
import json
from collections import defaultdict

def clean_price(val):
    if val is None: return None
    if isinstance(val, (int, float)): return float(val)
    if isinstance(val, str):
        s = re.sub(r'[^\d]', '', val)
        try: return float(s)
        except: return None
    return None

def get_base_name_and_grade(full_name):
    full_name = full_name.strip()
    grade = 1.0
    grade_name = "1 сорт"
    
    if "2 сорт" in full_name:
        grade = 0.5
        grade_name = "2 сорт"
    elif "Экспериментальная" in full_name or "Эксперементальная" in full_name:
        grade = 0.7
        grade_name = "Экспериментальная"
        
    # Remove grade from base name for price matching
    base_name = re.sub(r'2 сорт|Экспериментальная|Эксперементальная', '', full_name).strip()
    return base_name, grade, grade_name

def process_sklad2_and_prices(sklad_file, price_file):
    # 1. Extract Prices from Price List
    prices = {} # {NamePart: Price}
    try:
        wb_p = openpyxl.load_workbook(price_file, read_only=True, data_only=True)
        sheet_p = wb_p.active
        for row in sheet_p.iter_rows(min_row=1):
            name = str(row[1].value) if row[1].value else ""
            dims = str(row[3].value) if row[3].value else ""
            p_grey = clean_price(row[6].value)
            p_color = clean_price(row[8].value)
            
            key = f"{name} {dims}".strip().lower()
            if p_grey: prices[f"{key} серый"] = p_grey
            if p_color: prices[f"{key} цветной"] = p_color
    except Exception as e:
        print(f"Price Error: {e}")

    # 2. Process Sklad2
    inventory = defaultdict(float) # {FullName: TotalQty}
    try:
        wb_s = openpyxl.load_workbook(sklad_file, read_only=True, data_only=True)
        sheet_s = wb_s.active
        for row in sheet_s.iter_rows(min_row=1):
            name = row[0].value
            qty_raw = str(row[1].value) if row[1].value else "0"
            if not name: continue
            
            # Extract number from "123 шт" or "123.45 м2"
            qty_match = re.search(r'[\d,.]+', qty_raw)
            qty = float(qty_match.group().replace(',', '.')) if qty_match else 0.0
            inventory[name.strip()] += qty
    except Exception as e:
        print(f"Sklad Error: {e}")

    # 3. Combine and Generate Table
    final_table = []
    for full_name, qty in inventory.items():
        base_name, grade_mult, grade_label = get_base_name_and_grade(full_name)
        
        # Best matching for price
        base_price = 0.0
        match_key = base_name.lower()
        # Simple heuristic: if 'Серый' in name, use grey price, else maybe color?
        # For now, let's look for exact or fuzzy match in our price list
        for pk, pv in prices.items():
            if pk in match_key or match_key in pk:
                base_price = pv
                break
        
        calc_price = base_price * grade_mult
        final_table.append({
            "Name": full_name,
            "Grade": grade_label,
            "Qty": qty,
            "BasePrice": base_price,
            "FinalPrice": calc_price,
            "TotalValue": calc_price * qty
        })
    
    return final_table

# Run
results = process_sklad2_and_prices(r'c:\Users\Пользователь\Desktop\plittex-erp\склад2.xlsx', r'c:\Users\Пользователь\Desktop\plittex-erp\Price_март2026.xlsx')

# Output Markdown Table
print("| Название товара | Сорт | Кол-во | Базовая цена | Цена за ед. | Итого |")
print("| :--- | :--- | :--- | :--- | :--- | :--- |")
for item in results:
    print(f"| {item['Name']} | {item['Grade']} | {item['Qty']:.2f} | {item['BasePrice']:.2f} | {item['FinalPrice']:.2f} | {item['TotalValue']:.2f} |")
