import openpyxl

def peek_xlsx(file_path):
    try:
        wb = openpyxl.load_workbook(file_path, read_only=True, data_only=True)
        print(f"Sheets: {wb.sheetnames}")
        sheet = wb.active
        print(f"Active Sheet: {sheet.title}")
        
        for i, row in enumerate(sheet.iter_rows(max_row=50), 1):
            values = [cell.value for cell in row]
            print(f"Row {i}: {values}")
            
    except Exception as e:
        print(f"Error: {e}")

peek_xlsx(r'c:\Users\Пользователь\Desktop\plittex-erp\Price_март2026.xlsx')
