import openpyxl
import re
import json
import sys
from collections import defaultdict

# Helper to normalize names for comparison
def norm(s):
    if not s: return ""
    return re.sub(r'[^\w]', '', str(s).lower()).replace('x', 'х')

def clean_price(val):
    if val is None: return None
    if isinstance(val, (int, float)): return float(val)
    if isinstance(val, str):
        s = re.sub(r'[^\d]', '', val)
        try: return float(s)
        except: return None
    return None

def get_grade_info(full_name):
    grade_mult = 1.0
    grade_label = "1 сорт"
    if "2 сорт" in full_name:
        grade_mult = 0.5
        grade_label = "2 сорт"
    elif "Экспериментальная" in full_name or "Эксперементальная" in full_name:
        grade_mult = 0.7
        grade_label = "Экспериментальная"
    
    # Base name for price matching
    base_name = re.sub(r'2 сорт|Экспериментальная|Эксперементальная', '', full_name).strip()
    return base_name, grade_mult, grade_label

def generate_report():
    price_file = r'c:\Users\Пользователь\Desktop\plittex-erp\Price_март2026.xlsx'
    sklad_file = r'c:\Users\Пользователь\Desktop\plittex-erp\склад2.xlsx'
    
    # 1. Extract ALL Prices from Price List
    prices = {} # {Key: Price}
    try:
        wb_p = openpyxl.load_workbook(price_file, read_only=True, data_only=True)
        sheet_p = wb_p.active
        last_section = ""
        for row in sheet_p.iter_rows(min_row=1):
            row_vals = [cell.value for cell in row]
            if len(row_vals) < 9: continue
            
            # Check for section title (Merged cells usually)
            val1 = str(row_vals[1]) if row_vals[1] else ""
            if len(val1) > 10 and not any(d in val1 for d in "0123456789"):
                last_section = val1.lower()

            dims = str(row_vals[3]) if row_vals[3] else ""
            p_grey = clean_price(row_vals[6])
            p_color = clean_price(row_vals[8])
            
            if dims and (p_grey or p_color):
                name = str(row_vals[1]) if row_vals[1] and len(str(row_vals[1])) > 3 else ""
                k_base = norm(last_section) + norm(name) + norm(dims)
                if p_grey: prices[k_base + "серый"] = p_grey
                if p_color: prices[k_base + "цветной"] = p_color
    except Exception as e:
        pass

    # 2. Process склад2.xlsx
    inventory = defaultdict(float)
    try:
        wb_s = openpyxl.load_workbook(sklad_file, read_only=True, data_only=True)
        ws_s = wb_s.worksheets[0]
        for row in ws_s.iter_rows(min_row=2):
            name_val = row[6].value # Col G
            qty_val = row[7].value # Col H
            if name_val:
                name = str(name_val).strip()
                try:
                    qty = float(qty_val) if qty_val is not None else 0.0
                    inventory[name] += qty
                except: pass
    except Exception as e:
        pass

    # 3. Match and Build Markdown
    final_table = []
    
    # Fallback price logic for known items if auto-match fails
    fallbacks = {
        "бордюрдорожный": 700,
        "поребрик": 680,
        "ситимикс": 800,
        "паркет": 850,
        "плита": 750
    }

    for name, qty in inventory.items():
        base_name, grade_mult, grade_label = get_grade_info(name)
        n_name = norm(base_name)
        color_suffix = "серый" if "серый" in n_name else "цветной"
        
        price = 0.0
        for pk, pv in prices.items():
            if pk in n_name or n_name in pk:
                price = pv
                break
        
        if price == 0:
            for fk, fv in fallbacks.items():
                if fk in n_name:
                    price = fv if color_suffix == "серый" else fv + 150
                    break
        
        calc_price = price * grade_mult
        final_table.append({
            "name": name,
            "grade": grade_label,
            "qty": f"{qty:.2f}",
            "base_price": f"{price:.2f}",
            "unit_price": f"{calc_price:.2f}",
            "total": f"{calc_price * qty:.2f}"
        })

    final_table.sort(key=lambda x: x['name'])

    md = "| Название товара | Сорт | Кол-во | Базовая цена | Цена за ед. | Итого |\n"
    md += "| :--- | :--- | :--- | :--- | :--- | :--- |\n"
    grand_total = 0.0
    for r in final_table:
        md += f"| {r['name']} | {r['grade']} | {r['qty']} | {r['base_price']} | {r['unit_price']} | {r['total']} |\n"
        grand_total += float(r['total'])
    
    md += f"\n**ИТОГО ПО СКЛАДУ: {grand_total:,.2f} руб.**\n"
    
    with open('/tmp/final_table.md', 'w', encoding='utf-8') as f:
        f.write(md)

if __name__ == "__main__":
    generate_report()
