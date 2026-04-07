import openpyxl
import re
import json
import sys
from collections import defaultdict

sys.stdout.reconfigure(encoding='utf-8')

def clean_price(val):
    if val is None: return None
    if isinstance(val, (int, float)): return float(val)
    if isinstance(val, str):
        # Remove all spaces and non-digit characters
        s = re.sub(r'[^\d]', '', val)
        try: return float(s)
        except: return None
    return None

def get_base_name_and_grade(full_name):
    full_name = full_name.strip()
    grade = 1.0
    grade_label = "1 сорт"
    
    if "2 сорт" in full_name:
        grade = 0.5
        grade_label = "2 сорт"
    elif "Экспериментальная" in full_name or "Эксперементальная" in full_name:
        grade = 0.7
        grade_label = "Экспериментальная"
        
    base_name = re.sub(r'2 сорт|Экспериментальная|Эксперементальная', '', full_name).strip()
    return base_name, grade, grade_label

def process():
    price_file = r'c:\Users\Пользователь\Desktop\plittex-erp\Price_март2026.xlsx'
    sklad_file = r'c:\Users\Пользователь\Desktop\plittex-erp\склад2.xlsx'
    
    # 1. Harvest ALL possible Prices from Price List
    # We create a dictionary of {(Title, Dims, Color): Price}
    prices = {}
    try:
        wb_p = openpyxl.load_workbook(price_file, read_only=True, data_only=True)
        sheet_p = wb_p.active
        last_title = ""
        for row in sheet_p.iter_rows(min_row=1):
            row_vals = [cell.value for cell in row]
            if len(row_vals) < 10: continue
            
            # Title is in Col B (idx 1). 
            title_val = str(row_vals[1]) if row_vals[1] else ""
            if len(title_val) > 5 and not any(d in title_val for d in "0123456789"):
                last_title = title_val.strip().lower()
            
            # Dims in Col D (idx 3)
            dims = str(row_vals[3]) if row_vals[3] else ""
            
            # Prices: Col G (idx 6) = Grey, Col I (idx 8) = Color
            p_grey = clean_price(row_vals[6])
            p_color = clean_price(row_vals[8])
            
            # If we have dims and at least one price, we have a candidate
            if dims and re.search(r'\d', dims):
                d_key = dims.lower().replace(' ', '').replace('x', 'х')
                t_key = last_title.replace(' ', '')
                if p_grey: 
                    # Key pattern: TITLE_DIMS_GREY
                    prices[f"{t_key}_{d_key}_серый"] = p_grey
                if p_color:
                    prices[f"{t_key}_{d_key}_цветной"] = p_color
        print(f"DEBUG: Harvested {len(prices)} price combinations.")
    except Exception as e:
        print(f"Price Error: {e}")

    # 2. Process Sklad2
    inventory = defaultdict(float) 
    try:
        wb_s = openpyxl.load_workbook(sklad_file, read_only=True, data_only=True)
        sheet_s = wb_s.worksheets[0] 
        for row in sheet_s.iter_rows(min_row=2):
            name = row[6].value # Col G
            qty_val = row[7].value # Col H
            if not name: continue
            try:
                qty = float(qty_val) if qty_val is not None else 0.0
                inventory[str(name).strip()] += qty
            except: continue
        print(f"DEBUG: Merged into {len(inventory)} unique items.")
    except Exception as e:
        print(f"Sklad Error: {e}")

    # 3. Match and Price
    final_rows = []
    for full_name, qty in inventory.items():
        base_name, grade_mult, grade_label = get_base_name_and_grade(full_name)
        
        # Norm Name for matching
        norm_name = base_name.lower().replace(' ', '').replace('x', 'х')
        
        # Dimensional extraction from base_name if possible (needed for price lookup)
        dims_match = re.search(r'\d+х\d+х\d+', norm_name) or re.search(r'\d+х\d+', norm_name)
        dim_key = dims_match.group() if dims_match else ""
        
        # Color key
        color_key = "серый" if "серый" in norm_name else "цветной"
        
        price = 0.0
        # Search strategy: try to find a price where the harvested key bits match our current item
        for pk, pv in prices.items():
            # pk is like "бордюрдорожный_1000х300х150_серый"
            parts = pk.split('_')
            if len(parts) == 3:
                title_p, dims_p, color_p = parts
                if dims_p in norm_name and color_p in norm_name:
                    if title_p in norm_name or any(w in norm_name for w in title_p.split()):
                        price = pv
                        break
        
        # Special fallback for specific items if match failed
        if price == 0:
            if "бордюр дорожный" in base_name.lower():
                price = 700 if "серый" in base_name.lower() else 850
            elif "поребрик" in base_name.lower():
                price = 680
        
        calc_price = price * grade_mult
        final_rows.append([full_name, grade_label, f"{qty:.2f}", f"{price:.2f}", f"{calc_price:.2f}", f"{calc_price * qty:.2f}"])

    print("\n| Название товара | Сорт | Кол-во | Базовая цена | Цена за ед. | Итого |")
    print("| :--- | :--- | :--- | :--- | :--- | :--- |")
    for r in sorted(final_rows, key=lambda x: x[0]):
        print(f"| {' | '.join(r)} |")

process()
